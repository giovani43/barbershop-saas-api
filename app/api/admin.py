import re
import urllib.parse
from datetime import datetime, timezone, timedelta
from functools import wraps

import pytz
from flask import Blueprint, current_app, jsonify, request
from itsdangerous import BadSignature, SignatureExpired, URLSafeTimedSerializer
from sqlalchemy import text, func
from werkzeug.security import check_password_hash, generate_password_hash

from app.extensions import db
from app.models import Appointment, Barber, Client, Service, Shop

bp  = Blueprint("admin", __name__)
ART = pytz.timezone("America/Argentina/Buenos_Aires")

# ── Token helpers ──────────────────────────────────────────────────────────────

def _make_token(shop_id: str) -> str:
    s = URLSafeTimedSerializer(current_app.config["SECRET_KEY"])
    return s.dumps({"shop_id": shop_id}, salt="admin-v1")


def _verify_token(token: str):
    s = URLSafeTimedSerializer(current_app.config["SECRET_KEY"])
    try:
        data = s.loads(token, salt="admin-v1", max_age=86_400)
        return data.get("shop_id")
    except (BadSignature, SignatureExpired):
        return None


def admin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        auth     = request.headers.get("Authorization", "")
        token    = auth.removeprefix("Bearer ").strip()
        shop_id  = _verify_token(token)
        if not shop_id:
            return jsonify({"error": "No autorizado"}), 401
        shop = db.session.get(Shop, shop_id)
        if not shop:
            return jsonify({"error": "Barbería no encontrada"}), 404
        return f(shop, *args, **kwargs)
    return wrapper

# ── Registration ───────────────────────────────────────────────────────────────

@bp.post("/register")
def register():
    data     = request.get_json() or {}
    name     = data.get("name", "").strip()
    password = data.get("password", "").strip()
    plan     = data.get("plan", "solo")

    if not name or not password:
        return jsonify({"error": "Nombre y contraseña requeridos"}), 422

    # Auto-generate slug from name
    slug_base = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")
    slug      = slug_base
    counter   = 1
    while Shop.query.filter_by(slug=slug).first():
        slug = f"{slug_base}-{counter}"
        counter += 1

    shop = Shop(
        slug                = slug,
        name                = name,
        address             = data.get("address", ""),
        whatsapp            = data.get("whatsapp", ""),
        plan                = plan if plan in ("solo", "shop") else "solo",
        flash_promo_active  = False,
        admin_password_hash = generate_password_hash(password),
    )
    db.session.add(shop)
    db.session.commit()
    return jsonify({"token": _make_token(shop.id), "shop": shop.to_dict()}), 201

# ── Auth ───────────────────────────────────────────────────────────────────────

@bp.post("/login")
def login():
    data      = request.get_json() or {}
    shop_slug = data.get("shop_slug", "").strip()
    password  = data.get("password",  "").strip()

    shop = Shop.query.filter_by(slug=shop_slug).first()
    if not shop or not shop.admin_password_hash:
        return jsonify({"error": "Credenciales inválidas"}), 401
    if not check_password_hash(shop.admin_password_hash, password):
        return jsonify({"error": "Credenciales inválidas"}), 401

    return jsonify({"token": _make_token(shop.id), "shop": shop.to_dict()})

# ── Shop settings ──────────────────────────────────────────────────────────────

@bp.get("/shop")
@admin_required
def get_shop(shop):
    return jsonify(shop.to_dict())


@bp.put("/shop")
@admin_required
def update_shop(shop):
    data = request.get_json() or {}
    for field in ("name", "logo_url", "address", "whatsapp", "plan"):
        if field in data:
            setattr(shop, field, data[field])
    if data.get("password"):
        shop.admin_password_hash = generate_password_hash(data["password"])
    db.session.commit()
    return jsonify(shop.to_dict())


@bp.post("/promo")
@admin_required
def toggle_promo(shop):
    shop.flash_promo_active = not shop.flash_promo_active
    db.session.commit()
    return jsonify({"flash_promo_active": shop.flash_promo_active})

# ── Barbers ────────────────────────────────────────────────────────────────────

@bp.get("/barbers")
@admin_required
def list_barbers(shop):
    barbers = Barber.query.filter_by(shop_id=shop.id).order_by(Barber.name).all()
    return jsonify([b.to_dict() for b in barbers])


@bp.post("/barbers")
@admin_required
def create_barber(shop):
    data = request.get_json() or {}
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"error": "Nombre requerido"}), 422

    slug_base = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")
    slug      = slug_base
    counter   = 1
    while Barber.query.filter_by(slug=slug).first():
        slug = f"{slug_base}-{counter}"
        counter += 1

    barber = Barber(
        shop_id   = shop.id,
        name      = name,
        slug      = slug,
        shop_name = shop.name,
        shop_slug = shop.slug,
        photo_url = data.get("photo_url"),
        specialty = data.get("specialty"),
        bio       = data.get("bio"),
        whatsapp  = data.get("whatsapp"),
        instagram = data.get("instagram"),
        is_active = True,
    )
    db.session.add(barber)
    db.session.commit()
    return jsonify(barber.to_dict()), 201


@bp.put("/barbers/<barber_id>")
@admin_required
def update_barber(shop, barber_id):
    barber = Barber.query.filter_by(id=barber_id, shop_id=shop.id).first_or_404()
    data   = request.get_json() or {}
    for field in ("name", "photo_url", "specialty", "bio", "whatsapp", "instagram", "is_active"):
        if field in data:
            setattr(barber, field, data[field])
    if "slug" in data and data["slug"]:
        new_slug = data["slug"].strip().lower()
        conflict = Barber.query.filter(Barber.slug == new_slug, Barber.id != barber_id).first()
        if conflict:
            return jsonify({"error": "Ese slug ya está en uso"}), 409
        barber.slug = new_slug
    db.session.commit()
    return jsonify(barber.to_dict())


@bp.delete("/barbers/<barber_id>")
@admin_required
def delete_barber(shop, barber_id):
    barber = Barber.query.filter_by(id=barber_id, shop_id=shop.id).first_or_404()
    barber.is_active = False
    db.session.commit()
    return jsonify({"ok": True})


@bp.post("/barbers/<barber_id>/set-password")
@admin_required
def set_barber_password(shop, barber_id):
    barber   = Barber.query.filter_by(id=barber_id, shop_id=shop.id).first_or_404()
    data     = request.get_json() or {}
    password = data.get("password", "").strip()
    if len(password) < 4:
        return jsonify({"error": "La contraseña debe tener al menos 4 caracteres"}), 422
    barber.password_hash = generate_password_hash(password)
    db.session.commit()
    return jsonify({"ok": True})

# ── Appointment cleanup ────────────────────────────────────────────────────────

@bp.delete("/appointments/by-codes")
@admin_required
def delete_appointments_by_codes(shop):
    """Elimina turnos por booking_code (solo de esta shop)."""
    data  = request.get_json() or {}
    codes = data.get("codes", [])
    if not codes:
        return jsonify({"error": "Lista de códigos requerida"}), 422

    result = db.session.execute(text("""
        DELETE FROM appointments
        WHERE booking_code = ANY(:codes)
          AND barber_id IN (
              SELECT id FROM barbers WHERE shop_id = :shop_id
          )
        RETURNING booking_code
    """), {"codes": codes, "shop_id": shop.id})
    deleted = [r[0] for r in result]
    db.session.commit()
    return jsonify({"deleted": deleted, "count": len(deleted)})

# ── Services ───────────────────────────────────────────────────────────────────

@bp.get("/services")
@admin_required
def list_services(shop):
    services = (Service.query
                .filter_by(shop_id=shop.id)
                .order_by(Service.display_order)
                .all())
    return jsonify([s.to_dict() for s in services])


@bp.post("/services")
@admin_required
def create_service(shop):
    data  = request.get_json() or {}
    name  = data.get("name", "").strip()
    price = data.get("price")
    if not name or price is None:
        return jsonify({"error": "Nombre y precio requeridos"}), 422

    svc = Service(
        shop_id          = shop.id,
        name             = name,
        price            = float(price),
        duration_minutes = int(data.get("duration_minutes", 30)),
        display_order    = int(data.get("display_order", 0)),
        is_active        = True,
    )
    db.session.add(svc)
    db.session.commit()
    return jsonify(svc.to_dict()), 201


@bp.put("/services/<service_id>")
@admin_required
def update_service(shop, service_id):
    svc  = Service.query.filter_by(id=service_id, shop_id=shop.id).first_or_404()
    data = request.get_json() or {}
    for field in ("name", "duration_minutes", "display_order", "is_active"):
        if field in data:
            setattr(svc, field, data[field])
    if "price" in data:
        svc.price = float(data["price"])
    db.session.commit()
    return jsonify(svc.to_dict())


@bp.delete("/services/<service_id>")
@admin_required
def delete_service(shop, service_id):
    svc = Service.query.filter_by(id=service_id, shop_id=shop.id).first_or_404()
    svc.is_active = False
    db.session.commit()
    return jsonify({"ok": True})

# ── Stats / Agenda ─────────────────────────────────────────────────────────────

@bp.get("/stats")
@admin_required
def get_stats(shop):
    date_str = request.args.get("date", datetime.now(ART).strftime("%Y-%m-%d"))
    try:
        date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        return jsonify({"error": "Formato de fecha inválido"}), 400

    start_utc = ART.localize(datetime(date.year, date.month, date.day,  0,  0)).astimezone(timezone.utc)
    end_utc   = ART.localize(datetime(date.year, date.month, date.day, 23, 59)).astimezone(timezone.utc)

    barber_ids = [
        b.id for b in Barber.query.filter_by(shop_id=shop.id, is_active=True).all()
    ]
    if not barber_ids:
        return jsonify({"stats": {"booked": 0, "total": 0, "revenue": 0, "barbers": 0}, "agenda": []})

    ids_placeholder = ", ".join(f"'{bid}'" for bid in barber_ids)
    rows = db.session.execute(text(f"""
        SELECT
            a.id::text,
            b.name      AS barber_name,
            c.full_name AS client_name,
            c.whatsapp  AS client_wa,
            c.dni       AS client_dni,
            a.appointment_time,
            a.status,
            a.service_name,
            a.price,
            a.booking_code,
            a.qr_token,
            a.whatsapp_number,
            a.absence_charge_amount,
            a.absence_charge_sent,
            a.rescheduled_count
        FROM appointments a
        JOIN  barbers b ON b.id = a.barber_id
        LEFT JOIN clients c ON c.id = a.client_id
        WHERE a.barber_id IN ({ids_placeholder})
          AND a.appointment_time BETWEEN :start AND :end
        ORDER BY a.appointment_time
    """), {"start": start_utc, "end": end_utc}).mappings().all()

    agenda        = []
    booked_count  = 0
    total_revenue = 0.0
    late_minutes  = current_app.config.get("LATE_TOLERANCE_MINUTES", 8)
    charge_pct    = current_app.config.get("ABSENCE_CHARGE_PERCENT", 30)
    now_utc       = datetime.now(timezone.utc)

    for r in rows:
        appt_utc = r["appointment_time"]
        if appt_utc.tzinfo is None:
            appt_utc = appt_utc.replace(tzinfo=timezone.utc)
        local_t = appt_utc.astimezone(ART)

        if r["status"] in ("booked", "rescheduled"):
            booked_count  += 1
            total_revenue += float(r["price"]) if r["price"] else 0

        price_val   = float(r["price"]) if r["price"] else 0
        absence_fee = round(price_val * charge_pct / 100)

        # Botón "Cobrar ausencia": turno ya pasó + tolerancia y sigue en booked/rescheduled
        tolerance_dt   = appt_utc + timedelta(minutes=late_minutes)
        show_no_show   = (
            r["status"] in ("booked", "rescheduled")
            and now_utc >= tolerance_dt
        )

        agenda.append({
            "id":                   r["id"],
            "barber_name":          r["barber_name"],
            "client_name":          r["client_name"],
            "client_wa":            r["client_wa"],
            "client_dni":           r["client_dni"],
            "hora":                 local_t.strftime("%H:%M"),
            "fecha":                local_t.strftime("%d/%m/%Y"),
            "status":               r["status"],
            "service_name":         r["service_name"],
            "price":                price_val,
            "booking_code":         r["booking_code"],
            "qr_token":             r["qr_token"],
            "whatsapp_number":      r["whatsapp_number"],
            "absence_fee":          absence_fee,
            "absence_charge_amount": r["absence_charge_amount"] or 0,
            "absence_charge_sent":  bool(r["absence_charge_sent"]),
            "rescheduled_count":    r["rescheduled_count"] or 0,
            "show_no_show_btn":     show_no_show,
        })

    return jsonify({
        "stats": {
            "booked":  booked_count,
            "total":   len(rows),
            "revenue": total_revenue,
            "barbers": len(barber_ids),
        },
        "agenda": agenda,
    })


# ── POST /appointments/<id>/no-show ───────────────────────────────────────────

@bp.post("/appointments/<appointment_id>/no-show")
@admin_required
def mark_no_show(shop, appointment_id):
    """
    El barbero marca el turno como ausente.
    Calcula la multa (30% del servicio), la guarda y devuelve el link de WhatsApp.
    """
    appt = db.session.execute(text("""
        SELECT a.*, b.shop_id, c.full_name AS client_name, c.whatsapp AS client_wa
        FROM appointments a
        JOIN barbers b ON b.id = a.barber_id
        LEFT JOIN clients c ON c.id = a.client_id
        WHERE a.id = :id
    """), {"id": appointment_id}).mappings().first()

    if not appt:
        return jsonify({"error": "Turno no encontrado"}), 404
    if appt["shop_id"] != shop.id:
        return jsonify({"error": "No autorizado"}), 403
    if appt["status"] not in ("booked", "rescheduled"):
        return jsonify({"error": "Solo se puede cobrar ausencia en turnos reservados"}), 409

    price_val   = float(appt["price"]) if appt["price"] else 0
    charge_pct  = current_app.config.get("ABSENCE_CHARGE_PERCENT", 30)
    absence_fee = round(price_val * charge_pct / 100)

    # Formatear fecha/hora para el mensaje
    appt_utc = appt["appointment_time"]
    if appt_utc.tzinfo is None:
        appt_utc = appt_utc.replace(tzinfo=timezone.utc)
    ART_tz  = pytz.timezone("America/Argentina/Buenos_Aires")
    local_t = appt_utc.astimezone(ART_tz)
    fecha   = local_t.strftime("%d/%m/%Y")
    hora    = local_t.strftime("%H:%M")

    frontend_url = current_app.config.get("FRONTEND_URL", "")
    mp_alias     = current_app.config.get("MERCADOPAGO_ALIAS", "resquin.mvz")

    client_name = appt["client_name"] or "cliente"
    svc_name    = appt["service_name"] or "Servicio"

    wa_msg = (
        f"Hola {client_name} 👋, te contactamos desde MVZ Barbería.\n\n"
        f"Lamentablemente no te presentaste a tu turno del {fecha} a las {hora}hs "
        f"(o llegaste después de los 8 minutos de tolerancia permitidos).\n\n"
        f"De acuerdo a los Términos y Condiciones que aceptaste al reservar, "
        f"se aplica una multa del {charge_pct}% del valor del servicio como compensación "
        f"por el tiempo reservado:\n\n"
        f"💈 Servicio: {svc_name}\n"
        f"💰 Valor del servicio: ${int(price_val):,}\n"
        f"⚠️ Multa ({charge_pct}%): ${absence_fee:,}\n\n"
        f"Por favor realizá el pago al siguiente alias de Mercado Pago:\n"
        f"👉 {mp_alias}\n\n"
        f"Para más información sobre nuestra política visitá:\n"
        f"{frontend_url}/terminos"
    )

    wa_number = (appt["client_wa"] or appt["whatsapp_number"] or "").replace("+", "").replace(" ", "")
    wa_link   = f"https://wa.me/{wa_number}?text={urllib.parse.quote(wa_msg)}" if wa_number else None

    db.session.execute(text("""
        UPDATE appointments
        SET status                = 'no_show',
            absence_charge_amount = :fee,
            absence_charge_sent   = TRUE,
            updated_at            = NOW()
        WHERE id = :id
    """), {"fee": absence_fee, "id": appointment_id})

    db.session.commit()

    return jsonify({
        "message":      "Turno marcado como ausente.",
        "absence_fee":  absence_fee,
        "wa_link":      wa_link,
    }), 200

# ── Client import (XLSX) ───────────────────────────────────────────────────────

@bp.post("/clients/import")
@admin_required
def import_clients(shop):
    if "file" not in request.files:
        return jsonify({"error": "No se envió archivo"}), 400

    file = request.files["file"]
    if not file.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "Solo se aceptan archivos .xlsx"}), 400

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file.stream)
        ws = wb.active
    except Exception as e:
        return jsonify({"error": f"Error leyendo el archivo: {e}"}), 400

    headers = [
        str(cell.value).strip().lower() if cell.value else ""
        for cell in ws[1]
    ]

    try:
        idx_name = next(i for i, h in enumerate(headers) if "nombre" in h or "name" in h)
        idx_dni  = next(i for i, h in enumerate(headers) if "dni"    in h)
        idx_wa   = next(i for i, h in enumerate(headers)
                        if "whatsapp" in h or "celular" in h or "tel" in h)
    except StopIteration:
        return jsonify({"error": "El archivo debe tener columnas: Nombre, DNI, WhatsApp"}), 400

    first_barber      = Barber.query.filter_by(shop_id=shop.id, is_active=True).first()
    default_barber_id = first_barber.id if first_barber else None

    created = 0
    errors  = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            name = str(row[idx_name]).strip()  if row[idx_name] else ""
            dni  = str(row[idx_dni]).replace(".", "").strip() if row[idx_dni] else ""
            wa   = str(row[idx_wa]).strip()    if row[idx_wa]  else ""

            if not name or not dni:
                continue

            if not Client.query.filter_by(dni=dni, barber_id=default_barber_id).first():
                db.session.add(Client(
                    full_name=name, dni=dni, whatsapp=wa,
                    barber_id=default_barber_id,
                ))
                created += 1
        except Exception as e:
            errors.append(f"Fila {row_num}: {e}")

    db.session.commit()
    return jsonify({"created": created, "errors": errors})


# ── Temporary: force-run all migrations ───────────────────────────────────────
@bp.get("/run-migrations")
def run_migrations_endpoint():
    """Force-run all idempotent migrations. Safe to call multiple times."""
    from app import _run_migrations
    results = []
    from sqlalchemy import text as _text

    stmts = [
        "ALTER TABLE barbers      ADD COLUMN IF NOT EXISTS shop_id    VARCHAR(36)",
        "ALTER TABLE barbers      ADD COLUMN IF NOT EXISTS shop_name  VARCHAR(200)",
        "ALTER TABLE barbers      ADD COLUMN IF NOT EXISTS shop_slug  VARCHAR(100)",
        "ALTER TABLE barbers      ADD COLUMN IF NOT EXISTS photo_url  VARCHAR(500)",
        "ALTER TABLE barbers      ADD COLUMN IF NOT EXISTS instagram  VARCHAR(200)",
        "ALTER TABLE barbers      ADD COLUMN IF NOT EXISTS specialty  VARCHAR(200)",
        "ALTER TABLE barbers      ADD COLUMN IF NOT EXISTS bio        TEXT",
        "ALTER TABLE barbers      ADD COLUMN IF NOT EXISTS whatsapp   VARCHAR(50)",
        "ALTER TABLE barbers      ADD COLUMN IF NOT EXISTS is_active  BOOLEAN DEFAULT TRUE",
        "ALTER TABLE barbers      ADD COLUMN IF NOT EXISTS created_at TIMESTAMPTZ DEFAULT NOW()",
        "ALTER TABLE appointments ADD COLUMN IF NOT EXISTS service_id            VARCHAR(36)",
        "ALTER TABLE appointments ADD COLUMN IF NOT EXISTS cancelled_at          TIMESTAMPTZ",
        "ALTER TABLE appointments ADD COLUMN IF NOT EXISTS created_at            TIMESTAMPTZ DEFAULT NOW()",
        "ALTER TABLE appointments ADD COLUMN IF NOT EXISTS updated_at            TIMESTAMPTZ DEFAULT NOW()",
        "ALTER TABLE appointments ADD COLUMN IF NOT EXISTS whatsapp_number       VARCHAR(50)",
        "ALTER TABLE appointments ADD COLUMN IF NOT EXISTS qr_token              VARCHAR(255)",
        "ALTER TABLE appointments ADD COLUMN IF NOT EXISTS booking_code          VARCHAR(20)",
        "ALTER TABLE appointments ADD COLUMN IF NOT EXISTS rescheduled_count     INTEGER DEFAULT 0",
        "ALTER TABLE appointments ADD COLUMN IF NOT EXISTS absence_charge_sent   BOOLEAN DEFAULT FALSE",
        "ALTER TABLE appointments ADD COLUMN IF NOT EXISTS absence_charge_amount INTEGER DEFAULT 0",
        "ALTER TABLE appointments ADD COLUMN IF NOT EXISTS terms_accepted_at     TIMESTAMPTZ",
        "ALTER TABLE barbers      ADD COLUMN IF NOT EXISTS password_hash         VARCHAR(256)",
        """
        CREATE TABLE IF NOT EXISTS users (
            id         SERIAL PRIMARY KEY,
            dni        VARCHAR(20) UNIQUE NOT NULL,
            name       VARCHAR(100) NOT NULL,
            whatsapp   VARCHAR(20) NOT NULL,
            created_at TIMESTAMP DEFAULT NOW()
        )
        """,
        "ALTER TABLE appointments ADD COLUMN IF NOT EXISTS user_id INTEGER REFERENCES users(id)",
        "CREATE UNIQUE INDEX IF NOT EXISTS uq_appt_qr_token      ON appointments (qr_token)      WHERE qr_token IS NOT NULL",
        "CREATE UNIQUE INDEX IF NOT EXISTS uq_appt_booking_code  ON appointments (booking_code)  WHERE booking_code IS NOT NULL",
        """
        CREATE TABLE IF NOT EXISTS blocked_slots (
            id           SERIAL PRIMARY KEY,
            barber_id    VARCHAR(36) REFERENCES barbers(id) ON DELETE CASCADE,
            blocked_date DATE        NOT NULL,
            blocked_time TIME,
            all_day      BOOLEAN     NOT NULL DEFAULT FALSE,
            reason       VARCHAR(100),
            created_at   TIMESTAMP   DEFAULT NOW()
        )
        """,
        "CREATE INDEX IF NOT EXISTS idx_blocked_barber_date ON blocked_slots(barber_id, blocked_date)",
        "ALTER TABLE shops ADD COLUMN IF NOT EXISTS owner_email VARCHAR(200)",
    ]

    for sql in stmts:
        label = sql.strip().splitlines()[0][:80]
        try:
            db.session.execute(_text(sql))
            db.session.commit()
            results.append({"sql": label, "status": "ok"})
        except Exception as e:
            db.session.rollback()
            results.append({"sql": label, "status": "error", "detail": str(e)})

    return jsonify({"migrations": results})
